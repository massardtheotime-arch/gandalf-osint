import threading
import json
import os
import sys
import subprocess
import shutil
import re
import platform
import urllib.error
import urllib.request
import tempfile
import zipfile
import shlex
from pathlib import Path

import webview

try:
    import yt_dlp
except ImportError:
    pass

IS_WINDOWS     = platform.system() == "Windows"
IS_MACOS       = platform.system() == "Darwin"
PRORES_PROFILE = "1"
PRORES_EXT     = ".mov"
VERSION_FALLBACK = "dev"
RELEASE_API_URL = "https://api.github.com/repos/massardtheotime-arch/gandalf-osint/releases/latest"

# On Windows, hide console windows spawned by subprocess
SUBPROCESS_FLAGS = subprocess.CREATE_NO_WINDOW if IS_WINDOWS else 0


def parse_version(tag):
    match = re.match(r"^v?(\d+)(?:\.(\d+))?(?:\.(\d+))?$", str(tag or "").strip())
    if not match:
        return None
    return tuple(int(part or 0) for part in match.groups())


def is_newer_version(candidate, current):
    candidate_version = parse_version(candidate)
    current_version = parse_version(current)
    if not candidate_version or not current_version:
        return False
    return candidate_version > current_version


def select_release_asset(assets):
    candidates = assets or []
    if IS_WINDOWS:
        preferred_names = ("gandalfosint.exe",)
        extensions = (".exe",)
    elif IS_MACOS:
        preferred_names = ("gandalfosint.zip",)
        extensions = (".zip", ".dmg")
    else:
        preferred_names = ()
        extensions = ()
    for preferred in preferred_names:
        for asset in candidates:
            name = asset.get("name", "")
            if name.lower() == preferred and asset.get("browser_download_url"):
                return {
                    "name": name,
                    "url": asset["browser_download_url"],
                    "size": asset.get("size") or 0,
                }
    for ext in extensions:
        for asset in candidates:
            name = asset.get("name", "")
            if name.lower().endswith(ext) and asset.get("browser_download_url"):
                return {
                    "name": name,
                    "url": asset["browser_download_url"],
                    "size": asset.get("size") or 0,
                }
    return None


def current_app_bundle_path():
    if not IS_MACOS or not getattr(sys, "frozen", False):
        return None
    executable = Path(sys.executable).resolve()
    for path in (executable, *executable.parents):
        if path.suffix == ".app":
            return str(path)
    return None


def current_windows_exe_path():
    if not IS_WINDOWS or not getattr(sys, "frozen", False):
        return None
    executable = Path(sys.executable).resolve()
    if executable.suffix.lower() == ".exe":
        return str(executable)
    return None


def app_data_dir():
    if IS_WINDOWS:
        base = os.environ.get("LOCALAPPDATA") or os.path.expanduser("~\\AppData\\Local")
        path = os.path.join(base, "GandalfOSINT")
    elif IS_MACOS:
        path = os.path.expanduser("~/Library/Application Support/GandalfOSINT")
    else:
        path = os.path.expanduser("~/.gandalf-osint")
    os.makedirs(path, exist_ok=True)
    return path


def pending_update_file():
    return os.path.join(app_data_dir(), "pending_update.json")


def update_download_dir():
    path = os.path.join(app_data_dir(), "updates")
    os.makedirs(path, exist_ok=True)
    return path


def write_pending_update(update, path):
    payload = {
        "version": update.get("latest_version", ""),
        "asset_name": (update.get("asset") or {}).get("name", os.path.basename(path)),
        "path": path,
    }
    with open(pending_update_file(), "w", encoding="utf-8") as f:
        json.dump(payload, f)


def read_pending_update():
    try:
        with open(pending_update_file(), "r", encoding="utf-8") as f:
            return json.load(f)
    except (OSError, json.JSONDecodeError):
        return None


def pending_update_status():
    pending = read_pending_update()
    if not pending:
        return None
    version = pending.get("version", "")
    path = pending.get("path", "")
    if not path or not os.path.isfile(path):
        clear_pending_update()
        return None
    if not is_newer_version(version, APP_VERSION):
        clear_pending_update()
        return None
    return {
        "state": "ready",
        "latest_version": version,
        "name": pending.get("asset_name") or os.path.basename(path),
        "path": path,
    }


def clear_pending_update():
    try:
        os.remove(pending_update_file())
    except OSError:
        pass


def resource_path(name):
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, name)


def get_app_version():
    env_version = os.environ.get("GANDALF_VERSION", "").strip()
    if env_version:
        return env_version
    try:
        with open(resource_path("version.txt"), "r", encoding="utf-8") as f:
            version = f.read().strip()
        return version or VERSION_FALLBACK
    except OSError:
        pass
    try:
        repo_dir = os.path.dirname(os.path.abspath(__file__))
        version = subprocess.check_output(
            ["git", "describe", "--tags", "--exact-match", "HEAD"],
            cwd=repo_dir, stderr=subprocess.DEVNULL, text=True,
            creationflags=SUBPROCESS_FLAGS).strip()
        return version or VERSION_FALLBACK
    except Exception:
        return VERSION_FALLBACK


APP_VERSION = get_app_version()


def install_pending_update_on_startup():
    pending = pending_update_status()
    if not pending:
        return False
    destination = pending.get("path", "")
    if IS_MACOS and destination.lower().endswith(".zip"):
        started = start_macos_zip_install(destination)
    elif IS_WINDOWS and destination.lower().endswith(".exe"):
        started = start_windows_exe_install(destination)
    else:
        started = False
    if started:
        return True
    return False


def start_macos_zip_install(destination):
    current_app = current_app_bundle_path()
    if not current_app:
        return False
    try:
        extract_dir = tempfile.mkdtemp(prefix="gandalf-update-")
        with zipfile.ZipFile(destination) as zf:
            zf.extractall(extract_dir)
        new_app = None
        for root, dirs, _files in os.walk(extract_dir):
            for dirname in dirs:
                if dirname == "GandalfOSINT.app":
                    new_app = os.path.join(root, dirname)
                    break
            if new_app:
                break
        if not new_app:
            shutil.rmtree(extract_dir, ignore_errors=True)
            return False

        script_path = os.path.join(extract_dir, "install_update.sh")
        pid = os.getpid()
        script = f"""#!/bin/bash
set -e
APP_SRC={shlex.quote(new_app)}
APP_DST={shlex.quote(current_app)}
BACKUP="$APP_DST.previous-update"
PENDING={shlex.quote(pending_update_file())}
UPDATE_FILE={shlex.quote(destination)}
PID={pid}
while kill -0 "$PID" 2>/dev/null; do
  sleep 0.2
done
rm -rf "$BACKUP"
mv "$APP_DST" "$BACKUP"
if ditto "$APP_SRC" "$APP_DST"; then
  xattr -dr com.apple.quarantine "$APP_DST" 2>/dev/null || true
  rm -rf "$BACKUP"
  rm -f "$PENDING" "$UPDATE_FILE"
  open "$APP_DST"
  rm -rf {shlex.quote(extract_dir)}
else
  rm -rf "$APP_DST"
  mv "$BACKUP" "$APP_DST"
  open "$APP_DST"
  exit 1
fi
"""
        with open(script_path, "w", encoding="utf-8") as f:
            f.write(script)
        os.chmod(script_path, 0o755)
        subprocess.Popen(["/bin/bash", script_path], stdout=subprocess.DEVNULL,
                         stderr=subprocess.DEVNULL, start_new_session=True)
        return True
    except Exception:
        return False


def start_windows_exe_install(destination):
    current_exe = current_windows_exe_path()
    if not current_exe:
        return False
    try:
        temp_dir = tempfile.mkdtemp(prefix="gandalf-update-")
        script_path = os.path.join(temp_dir, "install_update.cmd")
        pid = os.getpid()
        script = f"""@echo off
set "APP_SRC={destination}"
set "APP_DST={current_exe}"
set "PENDING={pending_update_file()}"
set "PID={pid}"
set "COPIED="
:wait
tasklist /FI "PID eq %PID%" | find "%PID%" >nul
if not errorlevel 1 (
  timeout /t 1 /nobreak >nul
  goto wait
)
for /L %%i in (1,1,20) do (
  copy /Y "%APP_SRC%" "%APP_DST%" >nul 2>nul
  if not errorlevel 1 (
    set "COPIED=1"
    goto copied
  )
  timeout /t 1 /nobreak >nul
)
:copied
if defined COPIED (
  start "" "%APP_DST%"
  del "%APP_SRC%" >nul 2>nul
  del "%PENDING%" >nul 2>nul
) else (
  start "" "%APP_DST%"
)
rmdir /S /Q "{temp_dir}" >nul 2>nul
"""
        with open(script_path, "w", encoding="utf-8") as f:
            f.write(script)
        subprocess.Popen(["cmd", "/c", "start", "", script_path],
                         stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                         creationflags=SUBPROCESS_FLAGS)
        return True
    except Exception:
        return False


def find_deno():
    """Return path to deno binary — bundled inside .app first, then system."""
    # 1. Bundled binary (works on any Mac, no installation required)
    bundled = resource_path("deno")
    if os.path.isfile(bundled):
        try:
            os.chmod(bundled, 0o755)
        except OSError:
            pass
        return bundled
    # 2. System-installed deno (fallback for dev runs)
    if not IS_WINDOWS:
        for d in ["/opt/homebrew/bin", "/usr/local/bin",
                  os.path.expanduser("~/.deno/bin")]:
            candidate = os.path.join(d, "deno")
            if os.path.isfile(candidate):
                return candidate
    return None


# Make deno available to yt-dlp (needed to solve YouTube EJS anti-bot challenges)
_deno_bin  = find_deno()
_DENO_FOUND = _deno_bin is not None
_DENO_PATH  = _deno_bin or ""
if _deno_bin:
    _deno_dir = os.path.dirname(_deno_bin)
    if _deno_dir:
        os.environ["PATH"] = _deno_dir + ":" + os.environ.get("PATH", "")


def find_ffmpeg():
    # On Windows the bundled binary is ffmpeg.exe
    for name in ("ffmpeg.exe", "ffmpeg"):
        bundled = resource_path(name)
        if os.path.isfile(bundled):
            if sys.platform != "win32":
                os.chmod(bundled, 0o755)
            return bundled
    ff = shutil.which("ffmpeg")
    if ff:
        return ff
    if IS_WINDOWS:
        for p in [
            os.path.join(os.environ.get("LOCALAPPDATA", ""), "ffmpeg", "bin", "ffmpeg.exe"),
            os.path.join(os.environ.get("ProgramFiles", ""), "ffmpeg", "bin", "ffmpeg.exe"),
            r"C:\ffmpeg\bin\ffmpeg.exe",
        ]:
            if os.path.isfile(p):
                return p
    else:
        for p in ["/opt/homebrew/bin/ffmpeg", "/usr/local/bin/ffmpeg", "/usr/bin/ffmpeg"]:
            if os.path.isfile(p):
                return p
    return None


def find_ffprobe(ffmpeg_path):
    """Derive ffprobe path from ffmpeg path, handling .exe on Windows."""
    if not ffmpeg_path:
        return None
    d = os.path.dirname(ffmpeg_path)
    probe = "ffprobe.exe" if IS_WINDOWS else "ffprobe"
    candidate = os.path.join(d, probe)
    if os.path.isfile(candidate):
        return candidate
    return shutil.which("ffprobe")


def fmt_dur(secs):
    if not secs: return ""
    secs = int(secs)
    h, m, s = secs // 3600, (secs % 3600) // 60, secs % 60
    return f"{h}:{m:02d}:{s:02d}" if h else f"{m}:{s:02d}"


def fmt_views(n):
    if not n: return ""
    if n >= 1_000_000: return f"{n/1_000_000:.1f}M vues"
    if n >= 1_000:     return f"{n/1_000:.0f}K vues"
    return f"{n} vues"


def fmt_size(b):
    if not b: return ""
    if b >= 1_000_000_000: return f"{b/1e9:.1f} GB"
    return f"{b/1_000_000:.0f} MB"


class VideoInfo:
    def __init__(self, url, raw):
        self.url       = url
        self.title     = raw.get("title", url[:60])
        self.dur       = fmt_dur(raw.get("duration"))
        self.views     = fmt_views(raw.get("view_count"))
        self.channel   = raw.get("channel") or raw.get("uploader", "")
        self.thumb_url = raw.get("thumbnail", "")
        self.formats   = self._parse(raw.get("formats", []))
        self.sel       = 0

    def _parse(self, raw):
        out = []
        for h in [2160, 1440, 1080, 720, 480, 360, 240, 144]:
            vids = [f for f in raw
                    if f.get("height") == h
                    and f.get("vcodec", "none") not in ("none", None, "")]
            if vids:
                best = max(vids, key=lambda f: f.get("tbr") or 0)
                sz = best.get("filesize") or best.get("filesize_approx")
                out.append({"label": f"{h}p", "badge": "VIDÉO",
                            "spec": f"bestvideo[height<={h}]+bestaudio/bestvideo[height<={h}]/best[height<={h}]/best",
                            "size": fmt_size(sz), "audio": False})
        if not out:
            out.append({"label": "Meilleure", "badge": "VIDÉO",
                        "spec": "bestvideo+bestaudio/best", "size": "", "audio": False})
        audio = [f for f in raw
                 if f.get("vcodec", "none") in ("none", None, "")
                 and f.get("acodec", "none") not in ("none", None, "")]
        if audio:
            out.append({"label": "Audio MP3", "badge": "AUDIO",
                        "spec": "bestaudio/best", "size": "", "audio": True})
        return out

    def to_dict(self):
        return {"url": self.url, "title": self.title, "dur": self.dur,
                "views": self.views, "channel": self.channel,
                "thumb_url": self.thumb_url, "formats": self.formats, "sel": self.sel}


class Api:
    def __init__(self):
        self._window        = None
        self.output_dir     = os.path.expanduser("~/Downloads")
        self.ffmpeg_path    = find_ffmpeg()
        self._video_infos   = []
        self._xlsx_rows     = []
        self._latest_update = None
        self._update_checking = False
        self._update_downloading = False
        self.running        = False
        self.analysing      = False
        self._last_file     = None
        self.transcode_mode = "prores"
        self.xlsx_quality   = "bestvideo+bestaudio/best"
        self.cookies_file    = ""
        self.cookies_browser = ""   # "safari", "chrome", "firefox"

    # ── JS → Python ──────────────────────────────────────────────────────────

    def get_initial_state(self):
        import threading
        def _startup_log():
            import time; time.sleep(0.6)
            if _DENO_FOUND:
                self._emit("log", f"✓ deno : {_DENO_PATH}")
            else:
                self._emit("log", "⚠️  deno introuvable — YouTube risque d'échouer")
            if self.ffmpeg_path:
                self._emit("log", f"✓ ffmpeg : {self.ffmpeg_path}")
            else:
                self._emit("log", "⚠️  ffmpeg introuvable")
        threading.Thread(target=_startup_log, daemon=True).start()
        return {
            "output_dir":        self.output_dir,
            "ffmpeg_available":  bool(self.ffmpeg_path),
            "transcode_mode":    self.transcode_mode,
            "cookies_browser":   self.cookies_browser,
            "app_version":       APP_VERSION,
            "pending_update":    pending_update_status(),
        }

    def check_for_update(self, manual=False):
        if self._update_checking:
            return {"state": "checking"}
        self._update_checking = True
        self._emit("update_check_started", {"manual": bool(manual)})
        threading.Thread(target=self._check_update_worker, args=(bool(manual),), daemon=True).start()
        return {"state": "checking"}

    def download_update(self):
        if self._update_downloading:
            return {"state": "downloading"}
        pending = pending_update_status()
        if pending:
            self._emit("update_download_done", pending)
            return pending
        if not self._latest_update or not self._latest_update.get("asset"):
            return {"state": "error", "message": "Aucune mise a jour disponible."}
        self._update_downloading = True
        self._emit("update_download_started", self._latest_update)
        threading.Thread(target=self._download_update_worker, daemon=True).start()
        return {"state": "downloading"}

    def install_update_now(self):
        if self.running or self.analysing:
            return {"state": "busy", "message": "Terminez les telechargements avant d'installer la mise a jour."}
        pending = pending_update_status()
        if not pending:
            return {"state": "error", "message": "Aucune mise a jour prete a installer."}
        started = install_pending_update_on_startup()
        if not started:
            release_url = (self._latest_update or {}).get("release_url", "")
            return {
                "state": "error",
                "message": "Installation automatique impossible. Ouvrez la release pour installer manuellement.",
                "release_url": release_url,
            }
        self._emit("update_installing", pending)

        def _quit():
            import time
            time.sleep(0.25)
            os._exit(0)

        threading.Thread(target=_quit, daemon=True).start()
        return {"state": "installing"}

    def set_cookies_browser(self, browser):
        # browser = "safari" | "chrome" | "firefox" | ""
        self.cookies_browser = browser

    def analyse_urls(self, urls_text):
        if self.analysing or self.running:
            return
        urls = [u.strip() for u in urls_text.splitlines() if u.strip()]
        if not urls:
            return
        self._video_infos.clear()
        self.analysing = True
        threading.Thread(target=self._fetch_all, args=(urls,), daemon=True).start()

    def set_quality(self, card_idx, fmt_idx):
        card_idx, fmt_idx = int(card_idx), int(fmt_idx)
        if 0 <= card_idx < len(self._video_infos):
            self._video_infos[card_idx].sel = fmt_idx

    def set_transcode(self, mode):
        self.transcode_mode = mode

    def set_xlsx_quality(self, spec):
        self.xlsx_quality = spec

    def download_all(self):
        if self.running or not self._video_infos:
            return
        self.running = True
        self._emit("download_started")
        threading.Thread(target=self._batch, daemon=True).start()

    def import_xlsx(self):
        result = self._window.create_file_dialog(
            webview.OPEN_DIALOG,
            file_types=('Excel files (*.xlsx *.xls)', 'All files (*.*)')
        )
        if not result:
            return None
        path = result[0]
        try:
            import openpyxl
        except ImportError:
            return {"error": "openpyxl manquant — pip install openpyxl"}
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            header_row = None
            for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
                vals = [str(v).strip().lower() if v else "" for v in row]
                if "key" in vals and "link" in vals:
                    header_row = i
                    headers = [str(v).strip() if v else "" for v in row]
                    break
            if not header_row:
                return {"error": "En-têtes introuvables (key + link requis)"}
            norm = [h.lower() for h in headers]
            self._xlsx_rows = []
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                d = {norm[i]: row[i] for i in range(len(norm)) if i < len(row)}
                if d.get("link") and str(d["link"]).startswith("http"):
                    self._xlsx_rows.append(d)
            return {"filename": os.path.basename(path), "count": len(self._xlsx_rows)}
        except Exception as e:
            return {"error": str(e)}

    def download_xlsx(self):
        if not self._xlsx_rows or self.running:
            return
        self.running = True
        self._emit("download_started")
        threading.Thread(target=self._batch_xlsx, daemon=True).start()

    def choose_dir(self):
        result = self._window.create_file_dialog(
            webview.FOLDER_DIALOG, directory=self.output_dir)
        if result:
            self.output_dir = result[0]
            return result[0]
        return ""

    def open_link(self, url):
        if sys.platform == "win32":
            os.startfile(url)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", url])
        else:
            subprocess.Popen(["xdg-open", url])

    # ── Internal ─────────────────────────────────────────────────────────────

    def _emit(self, event, data=None):
        payload = json.dumps(data) if data is not None else "null"
        if self._window:
            self._window.evaluate_js(f"window.onEvent('{event}', {payload})")

    def _check_update_worker(self, manual):
        try:
            pending = pending_update_status()
            if pending:
                pending["manual"] = manual
                self._emit("update_check_result", pending)
                self._emit("update_download_done", pending)
                return
            request = urllib.request.Request(
                RELEASE_API_URL,
                headers={
                    "Accept": "application/vnd.github+json",
                    "User-Agent": f"GandalfOSINT/{APP_VERSION}",
                },
            )
            with urllib.request.urlopen(request, timeout=10) as response:
                release = json.loads(response.read().decode("utf-8"))

            latest_tag = release.get("tag_name", "")
            if not parse_version(latest_tag):
                result = {"state": "error", "message": "Version GitHub invalide."}
            elif not is_newer_version(latest_tag, APP_VERSION):
                self._latest_update = None
                result = {
                    "state": "up_to_date",
                    "current_version": APP_VERSION,
                    "latest_version": latest_tag,
                    "manual": manual,
                }
            else:
                asset = select_release_asset(release.get("assets", []))
                if not asset:
                    result = {
                        "state": "error",
                        "message": "Aucun installateur compatible trouve dans la release.",
                        "latest_version": latest_tag,
                        "release_url": release.get("html_url", ""),
                    }
                else:
                    result = {
                        "state": "available",
                        "current_version": APP_VERSION,
                        "latest_version": latest_tag,
                        "release_name": release.get("name") or latest_tag,
                        "release_url": release.get("html_url", ""),
                        "asset": asset,
                        "manual": manual,
                    }
                    self._latest_update = result
            self._emit("update_check_result", result)
            if result.get("state") == "available":
                self._emit("update_available", result)
                self.download_update()
        except urllib.error.URLError as e:
            self._emit("update_check_result", {
                "state": "error",
                "message": f"Impossible de contacter GitHub: {e.reason}",
                "manual": manual,
            })
        except Exception as e:
            self._emit("update_check_result", {
                "state": "error",
                "message": str(e),
                "manual": manual,
            })
        finally:
            self._update_checking = False

    def _download_update_worker(self):
        update = self._latest_update or {}
        asset = update.get("asset") or {}
        url = asset.get("url")
        name = asset.get("name") or "GandalfOSINT-update"
        destination = os.path.join(update_download_dir(), name)
        try:
            request = urllib.request.Request(
                url,
                headers={"User-Agent": f"GandalfOSINT/{APP_VERSION}"},
            )
            with urllib.request.urlopen(request, timeout=30) as response:
                total = int(response.headers.get("Content-Length") or asset.get("size") or 0)
                downloaded = 0
                with open(destination, "wb") as f:
                    while True:
                        chunk = response.read(1024 * 256)
                        if not chunk:
                            break
                        f.write(chunk)
                        downloaded += len(chunk)
                        pct = int(downloaded * 100 / total) if total else 0
                        self._emit("update_download_progress", {
                            "percent": pct,
                            "downloaded": downloaded,
                            "total": total,
                            "name": name,
                        })
            write_pending_update(update, destination)
            self._emit("update_download_done", {
                "state": "ready",
                "path": destination,
                "name": name,
                "latest_version": update.get("latest_version", ""),
            })
        except Exception as e:
            self._emit("update_download_done", {
                "state": "error",
                "message": str(e),
                "name": name,
            })
        finally:
            self._update_downloading = False

    def _cookie_opts(self):
        if not self.cookies_browser:
            return {}
        return {"cookiesfrombrowser": (self.cookies_browser,)}

    def _fetch_all(self, urls):
        cookie_opts = self._cookie_opts()
        self._emit("log", f"🍪 Cookies: {self.cookies_browser or 'aucun'}")
        ydl_opts = {"quiet": False, "no_warnings": False, "skip_download": True,
                    "remote_components": ["ejs:github"]}
        if self.ffmpeg_path:
            ydl_opts["ffmpeg_location"] = os.path.dirname(self.ffmpeg_path)
        ydl_opts.update(cookie_opts)
        for i, url in enumerate(urls):
            self._emit("status", f"Analyse [{i+1}/{len(urls)}] {url[:60]}")
            try:
                with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                    raw = ydl.extract_info(url, download=False)
                info = VideoInfo(url, raw)
                self._video_infos.append(info)
                self._emit("video_analysed", info.to_dict())
            except Exception as e:
                self._emit("log", f"✗ Erreur analyse : {e}")
        self.analysing = False
        self._emit("analyse_done", len(self._video_infos))
        self._emit("status", f"{len(self._video_infos)} vidéo(s) prête(s).")

    def _batch(self):
        total = len(self._video_infos)
        for n, info in enumerate(self._video_infos, 1):
            self._emit("status", f"[{n}/{total}] Téléchargement…")
            self._run(n, info, total)
        self._emit("status", f"Terminé — {total} fichier(s).")
        self._emit("progress", 100)
        self.running = False
        self._emit("download_done")

    def _run(self, n, info, total):
        self._last_file = None
        fmt = info.formats[info.sel]
        try:
            opts = self._ydl_opts(fmt["spec"], n)
            self._emit("log", f"[{n}/{total}] {info.title[:60]}")
            with yt_dlp.YoutubeDL(opts) as ydl:
                ydl.download([info.url])
            dl = self._last_file
            if not dl or not os.path.isfile(dl):
                raise FileNotFoundError("Fichier introuvable.")
            tc = self.transcode_mode
            if tc != "none" and not fmt["audio"] and self._has_video(dl):
                self._transcode(n, dl, total, tc)
            else:
                self._emit("log", f"[{n}/{total}] ✓ {os.path.basename(dl)}")
        except Exception as e:
            self._emit("log", f"[{n}/{total}] ✗ {e}")

    def _ydl_opts(self, spec, n):
        out = os.path.join(self.output_dir, "%(title)s_%(id)s.%(ext)s")
        pp  = []
        if spec == "bestaudio/best":
            pp = [{"key": "FFmpegExtractAudio", "preferredcodec": "mp3", "preferredquality": "192"}]
        opts = {
            "format": spec, "outtmpl": out, "postprocessors": pp,
            "progress_hooks":      [lambda d, _n=n: self._dl_hook(d, _n)],
            "postprocessor_hooks": [self._pp_hook],
            "quiet": True, "no_warnings": False, "merge_output_format": "mp4",
            "remote_components": ["ejs:github"],
        }
        if self.ffmpeg_path:
            opts["ffmpeg_location"] = os.path.dirname(self.ffmpeg_path)
        opts.update(self._cookie_opts())
        return opts

    def _dl_hook(self, d, n):
        if d["status"] == "downloading":
            try:
                pct = float(d.get("_percent_str", "0%").strip().replace("%", ""))
                self._emit("progress", pct * 0.7)
            except ValueError:
                pass
            self._emit("status",
                f"[{n}] {d.get('_percent_str','').strip()}  "
                f"{d.get('_speed_str','?')}  ETA {d.get('_eta_str','?')}")

    def _pp_hook(self, d):
        if d["status"] == "finished":
            fp = (d.get("info_dict") or {}).get("filepath") or d.get("filename")
            if fp:
                self._last_file = fp

    def _has_video(self, path):
        ffprobe = find_ffprobe(self.ffmpeg_path)
        if not ffprobe:
            return True
        try:
            out = subprocess.check_output(
                [ffprobe, "-v", "error", "-select_streams", "v:0",
                 "-show_entries", "stream=codec_type",
                 "-of", "default=noprint_wrappers=1:nokey=1", path],
                stderr=subprocess.DEVNULL, text=True,
                creationflags=SUBPROCESS_FLAGS)
            return "video" in out
        except Exception:
            return True

    def _get_duration(self, path):
        ffprobe = find_ffprobe(self.ffmpeg_path)
        if not ffprobe:
            return None
        try:
            out = subprocess.check_output(
                [ffprobe, "-v", "error", "-show_entries", "format=duration",
                 "-of", "default=noprint_wrappers=1:nokey=1", path],
                stderr=subprocess.DEVNULL, text=True,
                creationflags=SUBPROCESS_FLAGS)
            return float(out.strip())
        except Exception:
            return None

    def _transcode(self, n, src, total, mode="prores", final_name=None):
        folder = os.path.dirname(src)
        base   = os.path.splitext(src)[0]
        if mode == "prores":
            dst = (base + "_prores" + PRORES_EXT) if src.endswith(PRORES_EXT) else (base + PRORES_EXT)
            if final_name:
                dst = os.path.join(folder, final_name + PRORES_EXT)
            self._emit("log", f"[{n}/{total}] ProRes 422 LT → {os.path.basename(dst)}")
            cmd = [self.ffmpeg_path, "-y", "-i", src,
                   "-c:v", "prores_ks", "-profile:v", PRORES_PROFILE,
                   "-vendor", "apl0", "-bits_per_mb", "8000", "-pix_fmt", "yuv422p10le",
                   "-c:a", "pcm_s24le", dst]
        else:
            dst = base + "_premiere.mp4" if src.endswith(".mp4") else base + ".mp4"
            if final_name:
                dst = os.path.join(folder, final_name + ".mp4")
            self._emit("log", f"[{n}/{total}] MP4 H.264 → {os.path.basename(dst)}")
            cmd = [self.ffmpeg_path, "-y", "-i", src,
                   "-c:v", "libx264", "-preset", "slow", "-crf", "18",
                   "-pix_fmt", "yuv420p", "-c:a", "aac", "-b:a", "320k",
                   "-movflags", "+faststart", dst]

        if os.path.abspath(src) == os.path.abspath(dst):
            tmp = src + ".tmp_tc" + os.path.splitext(dst)[1]
            cmd[-1] = tmp
        else:
            tmp = None

        duration = self._get_duration(src)
        try:
            proc = subprocess.Popen(cmd, stderr=subprocess.PIPE,
                                    universal_newlines=True, bufsize=1,
                                    creationflags=SUBPROCESS_FLAGS)
            t_re = re.compile(r"time=(\d+):(\d+):(\d+)\.(\d+)")
            for line in proc.stderr:
                m = t_re.search(line)
                if m and duration:
                    h, mn, s, cs = int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))
                    pct = min((h*3600+mn*60+s+cs/100)/duration*100, 99)
                    self._emit("progress", 70 + pct * 0.3)
            proc.wait()
            if proc.returncode == 0:
                if tmp and os.path.isfile(tmp):
                    if os.path.isfile(dst): os.remove(dst)
                    os.rename(tmp, dst)
                try:
                    if os.path.abspath(src) != os.path.abspath(dst):
                        os.remove(src)
                except OSError:
                    pass
                lbl = "ProRes 422 LT" if mode == "prores" else "MP4 Premiere"
                self._emit("log", f"[{n}/{total}] ✓ {lbl} : {os.path.basename(dst)}")
            else:
                self._emit("log", f"[{n}/{total}] ✗ ffmpeg erreur {proc.returncode}")
        except FileNotFoundError:
            self._emit("log", "ffmpeg introuvable.")

    def _batch_xlsx(self):
        total = len(self._xlsx_rows)
        for n, row in enumerate(self._xlsx_rows, 1):
            url  = str(row.get("link", "")).strip()
            name = self._build_name(row)
            self._emit("status", f"[{n}/{total}] {name[:70]}")
            self._run_named(n, url, name, total)
        self._emit("status", f"Terminé — {total} fichier(s) (Excel).")
        self._emit("progress", 100)
        self.running = False
        self._emit("download_done")

    def _build_name(self, row):
        def s(v):
            if v is None: return ""
            return str(v).strip().replace("/", "-")
        return (f"{s(row.get('key'))}_{s(row.get('publication_date'))}_"
                f"{s(row.get('description'))}_{s(row.get('location'))}")

    def _run_named(self, n, url, custom_name, total):
        self._last_file = None
        is_audio = self.xlsx_quality == "bestaudio/best"
        try:
            out  = os.path.join(self.output_dir, f"{custom_name}.%(ext)s")
            pp   = [{"key": "FFmpegExtractAudio", "preferredcodec": "mp3",
                     "preferredquality": "192"}] if is_audio else []
            opts = {
                "format": self.xlsx_quality, "outtmpl": out, "postprocessors": pp,
                "progress_hooks":      [lambda d, _n=n: self._dl_hook(d, _n)],
                "postprocessor_hooks": [self._pp_hook],
                "quiet": True, "no_warnings": False,
                "merge_output_format": "mp4", "restrictfilenames": False,
            }
            if self.ffmpeg_path:
                opts["ffmpeg_location"] = os.path.dirname(self.ffmpeg_path)
            self._emit("log", f"[{n}/{total}] {url[:70]}")
            with yt_dlp.YoutubeDL(opts) as ydl:
                ydl.download([url])
            dl = self._last_file
            if not dl or not os.path.isfile(dl):
                raise FileNotFoundError("Fichier introuvable.")
            tc = self.transcode_mode
            if tc != "none" and not is_audio and self._has_video(dl):
                self._transcode(n, dl, total, tc, final_name=custom_name)
            else:
                self._emit("log", f"[{n}/{total}] ✓  {os.path.basename(dl)}")
        except Exception as e:
            self._emit("log", f"[{n}/{total}] ✗  {e}")


if __name__ == "__main__":
    if install_pending_update_on_startup():
        sys.exit(0)

    api = Api()
    html_path = resource_path("app.html")
    icon_path = resource_path("icon.icns")

    # Convert to a proper file:// URI (handles Windows drive letters & backslashes)
    from pathlib import Path
    html_uri = Path(html_path).as_uri()

    window = webview.create_window(
        "Gandalf OSINT",
        html_path,
        js_api=api,
        width=960,
        height=880,
        min_size=(700, 600),
        background_color="#0a0a0d",
    )
    api._window = window
    webview.start()
